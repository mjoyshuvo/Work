from django.db import transaction
from rest_framework import serializers, viewsets, status
from apps.common.models import UploadFiles, Banks
from apps.sales.models import DummyServicePurchased
from apps.api.Pagination import LargeResultsSetPagination, CustomResultsSetPagination
from rest_framework.response import Response
import openpyxl


class BankSerializer(serializers.ModelSerializer):
    class Meta:
        model = Banks
        fields = '__all__'


class BankViewSet(viewsets.ModelViewSet):
    serializer_class = BankSerializer
    pagination_class = CustomResultsSetPagination
    permission_classes = []
    model = Banks
    queryset = Banks.objects.all()


class UploadFilesSerializer(serializers.ModelSerializer):
    class Meta:
        model = UploadFiles
        fields = '__all__'


class UploadFilesViewSet(viewsets.ModelViewSet):
    serializer_class = UploadFilesSerializer
    pagination_class = LargeResultsSetPagination
    permission_classes = []
    model = UploadFiles
    queryset = UploadFiles.objects.all()

    def create(self, request, *args, **kwargs):
        with transaction.atomic():
            file_type = self.request.query_params.get('file_type')
            if file_type:
                request.data['file_type'] = file_type
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            uploaded_file = UploadFiles.objects.get(id=serializer.data['id'])
            uploaded_file_path = uploaded_file.file
            wb_obj = openpyxl.load_workbook(uploaded_file_path)

            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            for i in range(2, m_row + 1):
                product_name = sheet_obj.cell(row=i, column=1).value
                unit_type = sheet_obj.cell(row=i, column=2).value
                quantity = sheet_obj.cell(row=i, column=3).value
                per_unit_price = sheet_obj.cell(row=i, column=4).value
                vat_per_unit = sheet_obj.cell(row=i, column=5).value
                unit_total = per_unit_price * quantity
                vat_amount = quantity * vat_per_unit
                price_after_vat = unit_total + vat_amount
                DummyServicePurchased(product_name=product_name, unit_type=unit_type, quantity=quantity,
                                      per_unit_price=per_unit_price, unit_total=unit_total,
                                      vat_per_unit=vat_per_unit, vat_amount=vat_amount,
                                      price_after_vat=price_after_vat).save()

            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
